from openpyxl import load_workbook
from fugashi import Tagger
import unidic_lite
import os

os.environ["MECABRC"] = os.path.join(unidic_lite.DICDIR, "dicrc")
tagger = Tagger()

def to_hiragana_tokens(text):
    kana_map = str.maketrans(
        "アイウエオカキクケコサシスセソタチツテトナニヌネノハヒフヘホマミムメモヤユヨラリルレロワヲンヴー",
        "あいうえおかきくけこさしすせそたちつてとなにぬねのはひふへほまみむめもやゆよらりるれろわをんぶー"
    )
    tokens = []
    for m in tagger(text):
        raw = m.feature[7] if len(m.feature) > 7 and m.feature[7] not in (None, "*") else m.surface
        clean = str(raw).split("-")[0]
        hira = clean.translate(kana_map).lower()
        tokens.append(hira)
    return tokens

def token_match(input_text, target_text):
    input_tokens = to_hiragana_tokens(input_text)
    target_tokens = to_hiragana_tokens(target_text)
    return any(tok in target_tokens for tok in input_tokens)

def check_food(text):
    excel_file = "石山の規約情報.xlsx"
    sheets = ["不食品", "可食品"]
    wb = load_workbook(excel_file, data_only=True)

    normalized_input = text.strip()
    results = []

    for sheet_name in sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):
            b_val = row[1].value
            c_val = row[2].value

            if not b_val:
                continue

            if normalized_input.lower() == str(b_val).strip().lower():
                msg = f"✅ {sheet_name}に完全一致しました：{b_val}"
                if c_val:
                    msg += " → 備考あり。注意してください。"
                results.append(msg)
                return results

            if token_match(normalized_input, b_val):
                msg = f"🔍 {sheet_name}に部分一致しました：{b_val}"
                if c_val:
                    msg += " → 備考あり。注意してください。"
                results.append(msg)

    if not results:
        results.append("この食品は不明です。LINE、Slack等で連絡してください。[受付時間：8時～21時]")

    return results

